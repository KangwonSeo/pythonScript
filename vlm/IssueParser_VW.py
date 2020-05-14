from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Color
import re
import sys


SEARCHING_FILTER="http://vlm.lge.com/issue/issues/?jql=project%20%3D%20VWICASKPM%20AND%20resolution%20%3D%20Unresolved%20AND%20component%20in%20(%22System%20-%20reset%22%2C%20%22System%20-%20Reset%2FFreeze%22)"
driver = webdriver.Chrome('D:/tool/chromedriver_win32/chromedriver.exe')

def log_in():
    driver.get("https://vlmissue.lge.com/cuas/?os_destination=%2Fsecure%2FDashboard.jspa")
    id = input("input id : ")
    pwd = input("input pwd : ")
    otp = input("input otp : ")
    driver.implicitly_wait(5)
    driver.find_element_by_name('userId').send_keys(id)
    driver.find_element_by_name('password').send_keys(pwd)
    driver.implicitly_wait(1)
    driver.find_element_by_name('otpPassword').send_keys(otp)
    driver.find_element_by_xpath('//*[@id="btnLogin"]').click()
    driver.implicitly_wait(2)

def save_data():
    global issueList
    issueList = []
#get total data
    driver.get(SEARCHING_FILTER)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    totalCount = soup.find("span", {"class": "results-count-total results-count-link"}).get_text()
    r=int(totalCount)//50 + 1
    for i in range(r):
        driver.get(SEARCHING_FILTER+"&&startIndex="+str(i*50))
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        infoTable = soup.find("table", {"id": "issuetable"})
        for a in infoTable.find_all(id=re.compile('^issuerow')):
            tag = a.find_all(['a','time','span'])
            t=[]
            for j in range(len(tag)):
                t.append(tag[j].text)
            issueList.append(t)

    issueList = list(set(map(tuple, issueList)))

def parse_data():
    global created
    global remain
    global reject
    global reopen
    global backup
    created = 0
    remain = 0
    reject = 0
    reopen = 0

    backup = []
    today=datetime.today().date()
    for issue in issueList:
        if "VWICASKPM" in issue[2] :
            cd = datetime.strptime(issue[10], "%Y/%m/%d").date()
            if ((today - cd).days <= 7) :
                created += 1
            status=issue[8]
            summary=issue[3]
            if "Open" in status or "Reopen" in status or "In Progress" in status or "Response From Reporter" in status or "Inquire To Reporter" in status:
                if "PERF" not in summary:
                    if "Inquire To Reporter" in status:
                        reject += 1
                    if "Reopen" in status:
                        reopen += 1
                    remain += 1
                    backup.append(issue)

def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22

    row_num = 1
    ws.cell(row=row_num, column=1).value = "Key"
    ws.cell(row=row_num, column=2).value = "Summary"
    ws.cell(row=row_num, column=3).value = "Assignee"
    ws.cell(row=row_num, column=4).value = "Status"
    ws.cell(row=row_num, column=5).value = "Created"
    ws.cell(row=row_num, column=6).value = "Updated"

    for i in range(1, 7):
        ws.cell(row=row_num, column=i).fill = PatternFill(patternType='solid', fgColor=Color('7FFFD4'))

    for lt in backup:
        row_num += 1
        col_num = 0
        for skip_num in range(len(lt)):
            if skip_num in [0,1,5,6,7,9,11,13,14]:
                continue
            col_num += 1
            ws.cell(row=row_num, column=col_num).value = lt[skip_num]
            if "Inquire To Reporter" in lt[8]:
                ws.cell(row=row_num, column=col_num).fill = PatternFill(patternType='solid', fgColor=Color('696969'))

    wb.save("ticket_external_VW.xlsx")
    wb.close()

def export_chart():
    today = datetime.today().date()
    fd = "2019/01/01"
    cd = datetime.strptime(fd, "%Y/%m/%d").date()
    cweek = (today - cd).days // 7 + 1

    wb = load_workbook("ICAS-3_Stability_Issue.xlsx")
    ws = wb['Sheet1']

    for i in range(5):
        for j in range(11):
            ws.cell(row=j + 2, column=i + 3).value = ws.cell(row=j + 2, column=i + 4).value

    for i in range(1, 13):
        if i in [1, 8]:
            continue
        if i in [2, 9]:
            ws.cell(row=i, column=7).value = "(CW" + str(cweek) + ")"
        else:
            ws.cell(row=i, column=7).value = 0

    print(ws.cell(row=2, column=7).value)

    ws['G5'] = remain
    lresult = ws.cell(row=7, column=6).value
    ws['G6'] = lresult + created - remain
    ws['G7'] = remain
    ws['G10'] = created
    ws['G11'] = lresult + created - remain
    ws['G12'] = created - (lresult + created - remain)
    # make hidden matrix
    offset = 100
    for i in range(2, 8):
        for j in range(2, 8):
            ws.cell(row=offset + j, column=i).value = ws.cell(row=i, column=j).value

    chart = LineChart()
    chart.title = "Stability issue status"
    chart.style = 13
    data1 = Reference(ws, min_col=7, min_row=offset + 2, max_row=offset + 7)
    cats = Reference(ws, min_col=2, min_row=offset + 3, max_row=offset + 7)
    chart.add_data(data1, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I2")

    offset = 100
    for i in range(9, 13):
        for j in range(2, 8):
            ws.cell(row=offset + j, column=i).value = ws.cell(row=i, column=j).value

    chart = LineChart()
    chart.title = "Issue IN vs OUT flow"
    chart.style = 13
    data1 = Reference(ws, min_col=12, min_row=offset + 2, max_row=offset + 7)
    cats = Reference(ws, min_col=9, min_row=offset + 3, max_row=offset + 7)
    chart.add_data(data1, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I16")

    wb.save("ICAS-3_Stability_Issue_VW.xlsx")
    wb.close()

print(sys.argv[0])
last_week = int(sys.argv[1])

log_in()
save_data()
parse_data()
print("remained this week = ", remain)
print("created this week = ", created)
print("reopened ticket = ", reopen)
print("resolved ticket = ", created+last_week-remain)
print("reject ticket = ", reject)
export_excel()
#export_chart()
