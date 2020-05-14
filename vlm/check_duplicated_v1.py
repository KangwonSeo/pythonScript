from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Color
import re
import sys

BASEADDR="http://vlm.lge.com"
SEARCHING_FILTER=BASEADDR+"/issue/issues/?jql=project%20%3D%20VWICASKPM%20AND%20text%20~%20"
driver = webdriver.Chrome('D:/tool/chromedriver_win32/chromedriver.exe')
DCOMMENT="This ticket is come from"
SCOMMENT="This ticket is duplicated with"
MCOMMENT="The related changes have been"
STARTV="440"
url_list=[]
comment_list=[]
status_list=[]
fixv_list=[]
assign_list=[]
issuelist=[]
duplicate_list=[]
def parse_excel():
    print("parse start")
    global issuelist
    wb = load_workbook(sys.argv[1])
    ws = wb['Tabelle1']
    i=1
    while True:
        i += 1
        if ws.cell(row=i, column=1).value is not None:
            issuelist.append(ws.cell(row=i, column=1).value)
        else :
            break

def log_in():
    driver.get("https://vlmissue.lge.com/cuas/?os_destination=%2Fsecure%2FDashboard.jspa")
    id = input("input id : ")
    pwd = input("input pwd : ")
    otp = input("input otp : ")
    driver.implicitly_wait(5)
    driver.find_element_by_name('userId').send_keys(id)
    driver.find_element_by_name('password').send_keys(pwd)
    driver.implicitly_wait(1)
    driver.find_element_by_name('otpPassword').send_keys(otp)
    driver.find_element_by_xpath('//*[@id="btnLogin"]').click()
    driver.implicitly_wait(2)

def find_url():
    for kpm in issuelist:
        driver.get(SEARCHING_FILTER+"\""+str(kpm)+"\"")
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        url=soup.find("table", {"id": "issuetable"})
        if url is None :
            url_list.append("")
        else :
            url=url.find(id=re.compile('^issuerow')).find('a').get('href')
            url_list.append(BASEADDR+url)

def visit_urls():
    global comment_list
    global status_list
    global fixv_list
    global assign_list
    global duplicate_list
    for url in url_list:
        if len(url) == 0 :
            comment_list.append("")
            status_list.append("")
            fixv_list.append("")
            assign_list.append("")
            duplicate_list.append("")
            continue

        comments=[]
        last_comment=""
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

        #add last comment
        last_comment=""
        pticket="RC"
        for comment in soup.find("div", {"class":"issuePanelContainer"}).find_all(id=re.compile('^comment-')):
            a = comment.find("div", {"class":"action-body flooded"})
            if(a):
                for ptag in a.find_all("p"):
                    comments.append(ptag.get_text())
                temp_comment=("\n".join(comments))
                comments.clear()
                if SCOMMENT in temp_comment:
                    offset=temp_comment.find(SCOMMENT)
                    if len(temp_comment) >= offset+len(SCOMMENT) + 10 :
                        pticket = temp_comment[offset+len(SCOMMENT)+4:offset+len(SCOMMENT)+11]
                        break
                if MCOMMENT in temp_comment:
                    continue
                last_comment=temp_comment
        duplicate_list.append(pticket)
        comment_list.append(last_comment)

        #add status_list
        status=soup.find("ul",{"id":"issuedetails"}).find("li",{"class":"item item-right"}).find("span").get_text()
        status_list.append(status)
        #add fix_list
        fix_version=soup.find("ul", {"id":"issuedetails"}).find("span", {"id":"fixfor-val"}).find("span")
        if fix_version is not None :
            fix_version = fix_version.get_text()
        else :
            fix_version = ""        
        fixv_list.append(fix_version.strip())

        #add assign_list
        assignee=soup.find("li", {"class":"people-details"}).find("span", {"class":"aui-avatar-inner"}).find("img")
        if assignee is not None :
            assignee = assignee.get("alt")
        else :
            assignee = ""
        assign_list.append(assignee)

def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    row_num = 1
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 75
    ws.column_dimensions['F'].width = 13

    ws.cell(row=row_num, column=1).value = "KPM number"
    ws.cell(row=row_num, column=2).value = "Status"
    ws.cell(row=row_num, column=3).value = "Fix version"
    ws.cell(row=row_num, column=4).value = "VLM page"
    ws.cell(row=row_num, column=5).value = "comment"
    ws.cell(row=row_num, column=6).value = "Mother"

    for i in range(len(issuelist)):
        row_num += 1
        ws.cell(row=row_num, column=1).value = issuelist[i]
        ws.cell(row=row_num, column=2).value = status_list[i]
        ws.cell(row=row_num, column=3).value = fixv_list[i]
        ws.cell(row=row_num, column=4).value = url_list[i]
        ws.cell(row=row_num, column=5).value = comment_list[i]
        ws.cell(row=row_num, column=6).value = duplicate_list[i]

    wb.save("check_duplicated.xlsx")
    wb.close()

parse_excel()
log_in()
find_url()
visit_urls()
export_excel()
