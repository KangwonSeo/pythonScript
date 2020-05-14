from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Color
import re
import sys
import time

BASEADDR="http://vlm.lge.com"
SEARCHING_FILTER=BASEADDR+"/issue/issues/?jql=project%20%3D%20VWICASKPM%20AND%20component%20in%20(%22Android%20Auto%20-%20Reset%2FFreeze%22%2C%20%22AppConnect%20-%20Reset%2FFreeze%22%2C%20%22CarPlay%20-%20Reset%2FFreeze%22%2C%20%22HMI%20-%20Reset%2FFreeze%22%2C%20%22Mirror%20Link%20-%20Reset%2FFreeze%22%2C%20%22Navigation%20-%20Reset%2FFreeze%22%2C%20%22Speech%20-%20Reset%2FFreeze%22%2C%20%22System%20-%20reset%22%2C%20%22System%20-%20Reset%2FFreeze%22)"
driver = webdriver.Chrome('D:/tool/chromedriver_win32/chromedriver.exe')
DCOMMENT="This ticket is come from"
SCOMMENT="This ticket is duplicated with"
MCOMMENT="The related changes have been"
STARTV="440"
url_list=[]
kpm_list=[]
kpm_listp=[]
affectv_list=[]
title_list=[]
rep_list=[]
comment_list=[]
status_list=[]
fixv_list=[]
vlm_list=[]
assign_list=[]
duplicate_list=[]

def log_in():
    driver.get("https://vlmissue.lge.com/cuas/?os_destination=%2Fsecure%2FDashboard.jspa")
    id = input("input id : ")
    pwd = input("input pwd : ")
    otp = input("input otp : ")
    driver.implicitly_wait(5)
    driver.find_element_by_name('userId').send_keys(id)
    driver.find_element_by_name('password').send_keys(pwd)
    driver.implicitly_wait(1)
    driver.find_element_by_name('otpPassword').send_keys(otp)
    driver.find_element_by_xpath('//*[@id="btnLogin"]').click()
    driver.implicitly_wait(2)

def find_url():
    global url_list
    driver.get(SEARCHING_FILTER)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    totalCount = soup.find("span", {"class": "results-count-total results-count-link"}).get_text()
    r=int(totalCount)//50 + 1
    for i in range(r):
        driver.get(SEARCHING_FILTER+"&&startIndex="+str(i*50))
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        for a in soup.find("table", {"id": "issuetable"}).find_all(id=re.compile('^issuerow')):
            url=a.find('a').get('href')
            reporter=a.find("td", {"class":"reporter"}).find("a")
            if reporter is not None:
                reporter=reporter.get_text()
            else :
                reporter=""
            date=a.find("td", {"class":"created"}).find("time")
            if date is not None:
                date = date.get_text()
            else :
                date = ""
            status=a.find("td", {"class":"status"}).find("span")
            if status is not None:
                status = status.get_text()
            else :
                status = ""
            if date.find("2020") >= 0 and status.find("Inquire") < 0 :
                if reporter.find("ivi") >= 0 or reporter.find("SandeepGaggara") or reporter.find("SarahPleyer") or reporter.find("RaghavaVivek") or reporter.find("Bhaskar") or reporter.find("sanghyun.moon") >= 0 :
                    url_list.append(BASEADDR+url)

def visit_urls():
    global kpm_list
    global kpm_listp
    global affectv_list
    global title_list
    global rep_list
    global comment_list
    global status_list
    global fixv_list
    global vlm_list
    global assign_list
    global duplicate_list
    for url in url_list:
        dup_flag=0
        created_version=STARTV
        comments=[]
        last_comment=""
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

        #add KPM_list
        kpm_number = soup.find("div",{"class":"aui-page-header-main"}).find("h1",{"id":"summary-val"}).get_text()[-8:-1]

        #add affect_list
        affect_version=soup.find("ul", {"id":"issuedetails"}).find("span", {"id":"versions-val"}).find("span")
        if affect_version is not None :
            affect_version = affect_version.get_text()
        else :
            affect_version = ""

        #add title_list
        title=soup.find("div", {"class":"aui-page-header-main"}).find("h1",{"id":"summary-val"}).get_text()

        #add rep_list
        rep= soup.find("div",{"id":"customfieldmodule"}).find("ul",{"class":"property-list"}).find("div",{"id":"customfield_11245-val"}).get_text()
        rep_list.append(rep.strip())
        #add last comment
        last_comment=""
        for comment in soup.find("div", {"class":"issuePanelContainer"}).find_all(id=re.compile('^comment-')):
            a = comment.find("div", {"class":"action-body flooded"})
            if(a):
                for ptag in a.find_all("p"):
                    comments.append(ptag.get_text())
                temp_comment=("\n".join(comments))
                comments.clear()
                if SCOMMENT in temp_comment:
                    offset=temp_comment.find(SCOMMENT)
                    if len(temp_comment) >= offset+len(SCOMMENT) + 10 :
                        pticket = temp_comment[offset+len(SCOMMENT)+4:offset+len(SCOMMENT)+11]
                        duplicate_list.append([pticket, kpm_number.strip(), affect_version.strip(), title])
                        dup_flag = 1
                        break
                if MCOMMENT in temp_comment:
                    continue
                last_comment=temp_comment
        if dup_flag == 1 :
            continue

        comment_list.append(last_comment)
        kpm_list.append(kpm_number.strip())
        kpm_listp.append(kpm_number.strip())
        affectv_list.append(affect_version.strip())
        title_list.append(title)

        #add status_list
        status=soup.find("ul",{"id":"issuedetails"}).find("li",{"class":"item item-right"}).find("span").get_text()
        status_list.append(status)
        #add fix_list
        fix_version=soup.find("ul", {"id":"issuedetails"}).find("span", {"id":"fixfor-val"}).find("span")
        if fix_version is not None :
            fix_version = fix_version.get_text()
        else :
            fix_version = ""        
        fixv_list.append(fix_version.strip())
        
        vlm_list.append(url)
        #add assign_list
        assignee=soup.find("li", {"class":"people-details"}).find("span", {"class":"aui-avatar-inner"}).find("img")
        if assignee is not None :
            assignee = assignee.get("alt")
        else :
            assignee = ""
        assign_list.append(assignee)

def duplicate_work():
    for item in duplicate_list:
        if item[0] in kpm_list:
            idx=kpm_list.index(item[0])
            kpm_listp[idx] = kpm_listp[idx] + "\n" + item[1]
            affectv_list[idx] = affectv_list[idx] + "\n" + item[2]
            title_list[idx] = title_list[idx] + "\n" + item[3]

def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    row_num = 1
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 9
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 75
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 45
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 9
    ws.cell(row=row_num, column=1).value = "KPM number"
    ws.cell(row=row_num, column=2).value = "Affected"
    ws.cell(row=row_num, column=3).value = "Subject"
    ws.cell(row=row_num, column=4).value = "Frequency"
    ws.cell(row=row_num, column=5).value = "Root Cause"
    ws.cell(row=row_num, column=6).value = "Status"
    ws.cell(row=row_num, column=7).value = "Target"
    ws.cell(row=row_num, column=8).value = "Ticket Link"
    ws.cell(row=row_num, column=9).value = "Assign"
    for i in range(1, 10):
        ws.cell(row=row_num, column=i).fill = PatternFill(patternType='solid', fgColor=Color('7FFFD4'))
        ws.cell(row=row_num, column=i).border = thin_border
        ws.cell(row=row_num, column=i).alignment = Alignment(horizontal='center')
    for i in range(len(kpm_list)):
        row_num += 1
        ws.cell(row=row_num, column=1).value = kpm_listp[i]
        ws.cell(row=row_num, column=2).value = affectv_list[i]
        ws.cell(row=row_num, column=3).value = title_list[i]
        ws.cell(row=row_num, column=4).value = rep_list[i]
        ws.cell(row=row_num, column=5).value = comment_list[i]+"\n"
        ws.row_dimensions[row_num].height = 16.5 * comment_list[i].count('\n')
        ws.cell(row=row_num, column=6).value = status_list[i]
        ws.cell(row=row_num, column=7).value = fixv_list[i]
        ws.cell(row=row_num, column=8).value = vlm_list[i]
        ws.cell(row=row_num, column=9).value = assign_list[i]

        if "Resolved" in status_list[i] or "Closed" in status_list[i]:
            for i in range(1,10):
                ws.cell(row=row_num, column=i).fill = PatternFill(patternType='solid', fgColor=Color('339900'))
                ws.cell(row=row_num, column=i).border = thin_border
                ws.cell(row=row_num, column=i).alignment = Alignment(vertical='top')
                ws.cell(row=row_num, column=i).font = Font(size=9)
        else :
            for i in range(1,10):
                ws.cell(row=row_num, column=i).fill = PatternFill(patternType='solid', fgColor=Color('FFFFFF'))
                ws.cell(row=row_num, column=i).border = thin_border
                ws.cell(row=row_num, column=i).alignment = Alignment(vertical='top')

    wb.save("ICAS3_Stability_issue.xlsx")
    wb.close()

log_in()
find_url()
visit_urls()
duplicate_work()
export_excel()
