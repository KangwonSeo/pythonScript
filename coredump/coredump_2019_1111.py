import serial
import time
import signal
import threading
import sys
import msvcrt
import ctypes
import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color

baud = 115200
sleep_time = 10
statusFlag = "none"
bootDone = False
loginCount = 0

def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    row_num = 1
    ws.cell(row=row_num, column=1).value = "Process Name"
    ws.cell(row=row_num, column=2).value = "Restart works"
    ws.cell(row=row_num, column=3).value = "Coredump written"

    col_num = 1
    for j in range(len(ps_list)):
        row_num += 1
        ws.cell(row=row_num, column=col_num).value = ps_list[j]
        ws.cell(row=row_num, column=col_num+1).value = reset_result[j]
        ws.cell(row=row_num, column=col_num+2).value = coredump_result[j]

    wb.save("coredump_test.xlsx")
    wb.close()

def waitBoot():
    if bootDone:
        return True      
    time.sleep(sleep_time)
    
    return False
    
def parsing_data(data):
    global bootDone
    global reset_result
    global coredump_result

    tmp = ''.join(data)
    tmp = tmp.replace("\n", "")

    if statusFlag == "none":
        return
    
    if statusFlag == "bootWaiting":
        if tmp.find("AGL Boot Checked") >= 0:
            bootDone = True
    
    if statusFlag == "kill":
        if tmp.find("skip reboot command") >= 0:
            reset_result[i] = "O"

        if tmp.find("save coredump size") >= 0:
            coredump_result[i] = "O"


def readThread(ser):
    line = []
    while not exitThread:
        for c in ser.read():
            line.append(chr(c))
            if c == 10:
                parsing_data(line)
                del line[:]   

# main #
exitThread = False
a=['platformmon', 'weston','displaymgr','startscreen','carifservice','MIB_ServiceRegi','touch-app','vector-sip','mcu-interpreter','rvc-proxy']
b=['hmiroot.sh','servicemanager','persistence-app','hmiroot.sh','vtee','JAVA','carservice_app','viwiproxy','viwiproxy','avahi-daemon']
c=['rsi-cdn','esp','security-manage','amfsservice','socks-sandbox','trace-daemon','trace-system-se','socks-cdp','qseecomd']
d=['frcservice_app','infotainmentrec','ofonod','qseecomd','dlt-daemon','dbus-daemon','dlt-system','btapp','audiomgnt','radio','ecallservice_ap']
e=['media-mgr','waveplayer','RadioDataCtrl','sdcm','spi-service','onl-user-iaa','wlanmgr','xinetd','telnetd','nginx','MEBNotification','adbd']
f=['testiface','amsr_vector_fs_','launch_someipd.','someipd_posix','runNavi-al.sh','afm-user-daemon','alm','av.mainapp.mib3','vtee','tts']
g=['speech','mirrorlink','eth-diag','swup','androidauto','network','mdnsd','onl-service-mgt','systemsp_app','intentbroker_ap','SLU_main','iAP2d']
h=['CarPlayMain','internetradio','gemweb-app','avfsd','browserservice','webappmgr','dhd_eventd','dhd_watchdog_th','wpa_supplicant','ts-app','dhcpd','nginx']

ps_list=a+b+c+d+e+f+g+h
reset_result = []
coredump_result = []
ps_name = ""
print(sys.argv[1])
i=0

today = datetime.datetime.now()
nowDatetime = today.strftime('%Y-%m-%d_%H_%M')

ser = serial.Serial(sys.argv[1], baud, timeout=0)

pftclCmd = 'pfctl -d\n'
loginCmd = 'ssh -q -o StrictHostKeyChecking=no root@192.168.0.7\n'
resetCmd = '/usr/bin/mcu_reset.sh\n'
exitThread = False
thread = threading.Thread(target=readThread, args=(ser,))
thread.start()

for lt in ps_list:
    reset_result.append("X")
    coredump_result.append("X")

    ps_name = lt
    time.sleep(1)
    
    statusFlag = "bootWaiting"
    while True:
        if waitBoot():
            break

    print("boot finish")
    ser.write(pftclCmd.encode())

    time.sleep(1)

    statusFlag = "login"
    ser.write(loginCmd.encode())
    print("login finish")
    time.sleep(sleep_time)
    statusFlag = "kill"
    killCmd = 'pgrep ' + lt + '| head | xargs -IFILE kill -6 FILE\n'
    ser.write(killCmd.encode())
    time.sleep(sleep_time)

    statusFlag = "none"
    ser.write(resetCmd.encode())
    bootDone = False
    time.sleep(sleep_time)
    i += 1


exitThread = True
print ("___EXIT___")
ser.close()
time.sleep(2)
export_excel()
