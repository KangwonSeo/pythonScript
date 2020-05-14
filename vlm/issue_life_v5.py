from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Color
import re
import sys

BASEADDR="http://vlm.lge.com"
SEARCHING_FILTER=BASEADDR+"/issue/issues/?jql=project%20%3D%20VWICASKPM%20AND%20component%20in%20(%22Android%20Auto%20-%20Reset%2FFreeze%22%2C%20%22AppConnect%20-%20Reset%2FFreeze%22%2C%20%22CarPlay%20-%20Reset%2FFreeze%22%2C%20%22HMI%20-%20Reset%2FFreeze%22%2C%20%22Mirror%20Link%20-%20Reset%2FFreeze%22%2C%20%22Navigation%20-%20Reset%2FFreeze%22%2C%20%22Speech%20-%20Reset%2FFreeze%22%2C%20%22System%20-%20reset%22%2C%20%22System%20-%20Reset%2FFreeze%22)"
driver = webdriver.Chrome('D:/tool/chromedriver_win32/chromedriver.exe')
DCOMMENT="This ticket is come from"
SCOMMENT="This ticket is duplicated with"
STARTV="440"
url_list=[]
kpm_list=[]
created_list=[]
affectv_list=[]
fixv_list=[]
assign_list=[]
vlm_list=[]
safe_member = ["rene.pankratz", "young9.kim", "yohan25.kim", "junggo.choi", "junwoo8342.choi", "sanny.chun", "gijeong.shin", ]
def log_in():
    driver.get("https://vlmissue.lge.com/cuas/?os_destination=%2Fsecure%2FDashboard.jspa")
    id = input("input id : ")
    pwd = input("input pwd : ")
    otp = input("input otp : ")
    driver.implicitly_wait(5)
    driver.find_element_by_name('userId').send_keys(id)
    driver.find_element_by_name('password').send_keys(pwd)
    driver.implicitly_wait(1)
    driver.find_element_by_name('otpPassword').send_keys(otp)
    driver.find_element_by_xpath('//*[@id="btnLogin"]').click()
    driver.implicitly_wait(2)

def find_url():
    global url_list
    driver.get(SEARCHING_FILTER)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    totalCount = soup.find("span", {"class": "results-count-total results-count-link"}).get_text()
    r=int(totalCount)//50 + 1
    for i in range(r):
        driver.get(SEARCHING_FILTER+"&&startIndex="+str(i*50))
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        for a in soup.find("table", {"id": "issuetable"}).find_all(id=re.compile('^issuerow')):
            url=a.find('a').get('href')
            reporter=a.find("td", {"class":"reporter"}).find("a")
            if reporter is not None:
                reporter=reporter.get_text()
            else :
                reporter=""
            date=a.find("td", {"class":"created"}).find("time")
            if date is not None:
                date = date.get_text()
            else :
                date = ""
            status=a.find("td", {"class":"status"}).find("span")
            if status is not None:
                status = status.get_text()
            else :
                status = ""
            if date.find("2020") >= 0 and status.find("Inquire") < 0 :
                if reporter.find("ivi") >= 0 or reporter.find("SandeepGaggara") or reporter.find("SarahPleyer") or reporter.find("RaghavaVivek") or reporter.find("Bhaskar") or reporter.find("sanghyun.moon") >= 0 :
                    url_list.append(BASEADDR+url)

def visit_urls():
    global kpm_list
    global affectv_list
    global fixv_list
    global assign_list
    global created_list
    for url in url_list:
        skip_flag=0
        created_version=STARTV
        comments=[]
        driver.get(url)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        kpm_number = soup.find("div",{"class":"aui-page-header-main"}).find("h1",{"id":"summary-val"}).get_text()[-8:-1]
        issue_detail=soup.find("ul", {"id":"issuedetails"})
        affect_version=issue_detail.find("span", {"id":"versions-val"}).find("span")
        assignee=soup.find("li", {"class":"people-details"}).find("span", {"class":"aui-avatar-inner"}).find("img")

        for comment in soup.find("div", {"class":"issuePanelContainer"}).find_all(id=re.compile('^comment')):
            a = comment.find("div", {"class":"action-body flooded"})
            if(a):
                for ptag in a.find_all("p"):
                    comments.append(ptag.get_text())
        for str in comments:
            if SCOMMENT in str:
                skip_flag = 1
                break
            if DCOMMENT in str:
                offset=str.find(DCOMMENT)
                if len(str) >= offset + len(DCOMMENT) + 5 :
                    created_version=str[offset+len(DCOMMENT)+2:offset+len(DCOMMENT)+5]
        if skip_flag == 1 :
            continue
        if affect_version is not None :
            affect_version = affect_version.get_text()
        else :
            affect_version = ""
        fix_version=issue_detail.find("span", {"id":"fixfor-val"}).find("span")
        if fix_version is not None :
            fix_version = fix_version.get_text()
        else :
            fix_version = ""
        if assignee is not None :
            assignee = assignee.get("alt")
        else :
            assignee = ""
        kpm_list.append(kpm_number)
        affectv_list.append(affect_version)
        fixv_list.append(fix_version)
        assign_list.append(assignee)
        created_list.append(created_version)
        vlm_list.append(url)

def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    row_num = 1
    ws.cell(row=row_num, column=1).value = ""
    ws.cell(row=row_num, column=2).value = "Created"
    ws.cell(row=row_num, column=3).value = "Detected"
    ws.cell(row=row_num, column=4).value = "Fixed"

    ws.cell(row=row_num, column=6).value = "Partition"
    ws.cell(row=row_num, column=7).value = "Assign"
    ws.cell(row=row_num, column=8).value = "KPM"
    for i in range(len(kpm_list)):
        row_num += 1
        if kpm_list[i].isdigit():
            ws.cell(row=row_num, column=1).value = int(kpm_list[i])
        else :
            ws.cell(row=row_num, column=1).value = kpm_list[i]
        if len(affectv_list[i]) > 3 :
            if affectv_list[i][-4:].isdigit():
                ws.cell(row=row_num, column=3).value = int(affectv_list[i][-4:])
                ws.cell(row=row_num, column=2).value = int(created_list[i]) if int(affectv_list[i][-4:]) >= int(created_list[i]) else int(affectv_list[i][-4:])
            else :
                ws.cell(row=row_num, column=3).value = affectv_list[i][-4:]
                ws.cell(row=row_num, column=2).value = int(STARTV)

        if len(fixv_list[i]) > 3 and "None" not in fixv_list[i] :
            if fixv_list[i][-4:].isdigit():
                ws.cell(row=row_num, column=4).value = int(fixv_list[i][-4:])
            else :
                ws.cell(row=row_num, column=4).value = fixv_list[i][-4:]
        ws.cell(row=row_num, column=7).value = assign_list[i]
        if assign_list[i] in safe_member:
            ws.cell(row=row_num, column=6).value = "safe"
        else:
            ws.cell(row=row_num, column=6).value = "IVI"
            
        ws.cell(row=row_num, column=8).value = vlm_list[i]
    wb.save("daily_issuelife.xlsx")
    wb.close()

log_in()
find_url()
visit_urls()
export_excel()
